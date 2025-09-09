from dataclasses import dataclass
from pathlib import Path
from typing import Protocol, Iterable, Mapping, Any, Optional, Sequence, Union, Callable, Dict
from openpyxl import load_workbook
import unicodedata

RowFilter = Callable[[Mapping[str, Any]], bool]

def normalize_header_text(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(character for character in value if not unicodedata.combining(character))
    value = " ".join(value.split())
    return value

class RecordReader(Protocol):
    def iterate_records(
        self,
        canonical_names: Sequence[str],
        row_filter: Optional[RowFilter] = None,
        max_records: Optional[int] = None,
    ) -> Iterable[Mapping[str, Any]]:
        ...

@dataclass
class ExcelStreamReader:
    path: Union[str, Path]
    sheet: Union[str, int] = 0

    def iterate_records(
        self,
        canonical_names: Sequence[str],
        row_filter: Optional[RowFilter] = None,
        max_records: Optional[int] = None,
    ) -> Iterable[Mapping[str, Any]]:
        workbook = load_workbook(filename=self.path, read_only=True, data_only=True)
        try:
            worksheet = workbook[self.sheet] if isinstance(self.sheet, str) else workbook.worksheets[self.sheet]

            header_row_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            normalized_headers_by_index: Dict[int, str] = {
                column_index: normalize_header_text(header_value)
                for column_index, header_value in enumerate(header_row_values)
            }

            requested_tokens = [normalize_header_text(name) for name in canonical_names]

            column_index_to_canonical_key: Dict[int, str] = {}
            matched_tokens: set[str] = set()

            for token, canonical_key in zip(requested_tokens, canonical_names):
                if not token or token in matched_tokens:
                    continue
                matched_index: Optional[int] = next(
                    (
                        column_index
                        for column_index, header_text in normalized_headers_by_index.items()
                        if header_text and token in header_text
                    ),
                    None,
                )
                if matched_index is not None:
                    column_index_to_canonical_key[matched_index] = normalize_header_text(canonical_key)
                    matched_tokens.add(token)

            yielded_count = 0
            for data_row_values in worksheet.iter_rows(min_row=2, values_only=True):
                record: Dict[str, Any] = {}
                total_cells_in_row = len(data_row_values)

                for column_index, canonical_key in column_index_to_canonical_key.items():
                    cell_value = data_row_values[column_index] if column_index < total_cells_in_row else None
                    record[canonical_key] = cell_value

                if row_filter is None or row_filter(record):
                    yield record
                    yielded_count += 1
                    if max_records is not None and yielded_count >= max_records:
                        break
        finally:
            workbook.close()
